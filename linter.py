from telegram import Update , ReplyKeyboardMarkup ,MessageEntity
from telegram.ext import ApplicationBuilder, ConversationHandler , CommandHandler , MessageHandler , ContextTypes,filters
from openpyxl import load_workbook
import pandas as pd
from random import choice
from datetime import datetime , timedelta

menu = [['review','find word'],['cancel']]
markup = ReplyKeyboardMarkup(menu,one_time_keyboard=True)
path = 'linter/linter.csv'
menu,findWord,addOrNot,addWord,isTrue,_continue = range(6)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global sheet , review_list , user
    sheet = pd.read_csv(path)
    user = 'p' if update.message.from_user['username'] == 'Cevat8809' else 'n'
    review_list = sheet[sheet[f'{user}date'] <= str(datetime.today().date())].kalame.to_list()
    name = update.message.from_user['first_name']
    Entity = [{'offset':6,'length':len(name),'type':'bold'}]
    await context.bot.send_sticker(chat_id=update.effective_chat.id, sticker='CAACAgQAAxkBAAEG_BhjpwzaE0_WGWvzgT9r-BK0Yy60GQACiwgAAmfHiVJJMa5s69pZ0iwE')
    await context.bot.send_message(chat_id= update.effective_chat.id, text=f'salam {name} joon khosh umadi\nomidvaram emruz kalamehaye khubi yad begirim',reply_markup=markup)
    return menu

async def get_word(updat:Update, context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=updat.effective_chat.id,text='lotfan kalameye mored nazar ro beman begu')
    return findWord

async def find_word(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global word
    word = update.message.text.lower()
    if not sheet[sheet.kalame == word].empty:
        await context.bot.send_message(chat_id=update.effective_chat.id,text=f'in kalame hast\n{sheet[sheet.kalame == word].mani.item()}')
        return menu
    else:
        await context.bot.send_sticker(chat_id=update.effective_chat.id,sticker='CAACAgQAAxkBAAEG_BJjpwlt0zHspfumsnhjHC2mTjvOjQACBAoAAvZRiVL44lglOLvBKCwE')
        await context.bot.send_message(chat_id=update.effective_chat.id,text='in kalame tuye dictionary nist.\nmikhay ezafe konim?',reply_markup=ReplyKeyboardMarkup([['YES','NO']]))
        return addOrNot

async def Error(update:Update,context:ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=update.effective_chat.id,text='dastor ya linki ke behem midi eshtebahe')
    return menu

async def start_review(update: Update,context: ContextTypes.DEFAULT_TYPE):
    global review_word
    if review_list == []:
        await context.bot.send_message(chat_id=update.effective_chat.id,text='kalamei baraye morur nist')
        return menu
    else:
        review_word = choice(review_list)
        await context.bot.send_message(chat_id=update.effective_chat.id,text=review_word)
        mani = sheet[sheet.kalame == review_word].mani.item()
        await context.bot.send_message(chat_id=update.effective_chat.id,text= mani ,entities=[{'offset':0,'length':len(mani),'type':'spoiler'}])
        await context.bot.send_message(chat_id=update.effective_chat.id,text='balad budi?',reply_markup=ReplyKeyboardMarkup([['yes','no']],one_time_keyboard=True))
    return isTrue

async def get_mean(updat: Update, context:ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=updat.effective_chat.id,text='hala ke mikhay ezafe koni lotfan mani kalame ro be man bede')
    return addWord

async def add_word(update: Update,context:ContextTypes.DEFAULT_TYPE):
    mean = update.message.text.lower()
    # wb = load_workbook(path)
    # sheet = wb.active
    # max_row = sheet.max_row
    # sheet[f'A{max_row}'] = word
    # sheet[f'B{max_row}'] = mean    
    # wb.save(path)
    sheet.loc[len(sheet.index)] = [word,mean,1,datetime.today().date(),1,datetime.today().date()]
    sheet.to_csv(path,index=False)
    await context.bot.send_message(chat_id=update.effective_chat.id,text='kalameye jadid save shod\nbargashtim menuye asli hala chikar konim?????',reply_markup=markup)
    return menu

async def return_menu(update:Update,context:ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=update.effective_chat.id,text='bargashtim menuye asli hala chikar konim?????',reply_markup=markup)
    return menu

async def check_review(update:Update,context:ContextTypes.DEFAULT_TYPE):
    a = update.message.text
    if a == 'yes':
        status_list = [1,3,5,7,15,30]
        i = sheet.index[sheet.kalame == review_word].item()
        old_status = sheet[sheet.index == i][f'{user}status'].item()
        sheet.at[i,f'{user}date'] =  datetime.today() + timedelta(days= old_status)
        sheet.at[i,f'{user}status'] = status_list[status_list.index(old_status)+1]
        sheet.to_csv(path,index=False)
    await context.bot.send_message(chat_id=update.effective_chat.id,text='mikhay edame bedi?',reply_markup=ReplyKeyboardMarkup([['YES','NO']],one_time_keyboard=True))
    return _continue

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=update.effective_chat.id,text='mersi ke az man estefade kardi\nharvaght khasti \'/start\' ro bezani man dar khedmatam')
    await context.bot.send_sticker(chat_id=update.effective_chat.id, sticker='CAACAgQAAxkBAAEHUw5jyA6HNrwCBpQrohvsAqYwPt4qEwAChwoAAqkbiVIwObBnZtRl8S0E')
    return ConversationHandler.END

def main():
    application = ApplicationBuilder().token('5024895126:AAGbQ7OBTxoiMxtdPtwATDNSzgY6Qlvcj7o').build()
    conv_handler = ConversationHandler(
        entry_points= [CommandHandler('start',start)],
        states= {
            menu : [MessageHandler(filters.Regex('^review$'),start_review),
                    MessageHandler(filters.Regex('^find word$'),get_word)],
            findWord: [MessageHandler(filters.Regex('.*'),find_word)],
            addOrNot: [MessageHandler(filters.Regex('^YES$'),get_mean),
                       MessageHandler(filters.Regex('^NO$'),return_menu)],
            addWord: [MessageHandler(filters.Regex('.*'),add_word)],
            isTrue: [MessageHandler(filters.Regex('.*'),check_review)],
            _continue: [MessageHandler(filters.Regex('^YES$'),start_review),
                       MessageHandler(filters.Regex('^NO$'),return_menu)]
        },
        fallbacks= [MessageHandler(filters.Regex("^cancel$"), cancel)]
    )
    application.add_handler(conv_handler)
    application.run_polling()

if __name__ == '__main__':
    main()